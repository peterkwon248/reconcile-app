import streamlit as st 
import pandas as pd
import io
from openpyxl.styles import PatternFill, Font
from openpyxl import load_workbook

st.set_page_config(page_title="📊 계산서 자동 정리 프로그램", layout="centered")
st.title("📊 계산서 자동 정리 프로그램")

st.markdown("### 📄 사이트 주문내역 엑셀 업로드")
order_file = st.file_uploader("", type=["xls", "xlsx"], key="order", label_visibility="collapsed")

st.markdown("### 💰 계좌 입금내역 엑셀 업로드")
deposit_file = st.file_uploader("", type=["xls", "xlsx"], key="deposit", label_visibility="collapsed")

if order_file and deposit_file:
    try:
        # ✅ 주문내역 처리
        order_df = pd.read_excel(order_file, engine="openpyxl")
        order_columns = order_df.columns

        # 자동 컬럼 이름 매핑
        order_df = order_df.rename(columns={
            [col for col in order_columns if '입금자' in col][0]: "입금자(사이트)",
            [col for col in order_columns if '주문자' in col or '회원명' in col][0]: "주문자",
            [col for col in order_columns if '결제' in col or '구매금액' in col][0]: "총 구매금액"
        })

        order_df["총 구매금액"] = pd.to_numeric(order_df["총 구매금액"], errors="coerce").fillna(0)
        order_df["입금자키"] = order_df["입금자(사이트)"].astype(str).str.replace(" ", "").str.strip()

        order_grouped = order_df.groupby("입금자키", as_index=False).agg({
            "주문자": "first",
            "입금자(사이트)": "first",
            "총 구매금액": "sum"
        })

        # ✅ 입금내역 처리
        deposit_df = pd.read_excel(deposit_file, engine="openpyxl")
        deposit_columns = deposit_df.columns

        deposit_df = deposit_df.rename(columns={
            [col for col in deposit_columns if '내용' in col or '입금자' in col][0]: "입금자(실제)",
            [col for col in deposit_columns if '금액' in col][0]: "통장입금"
        })

        deposit_df["통장입금"] = pd.to_numeric(deposit_df["통장입금"], errors="coerce").fillna(0)
        deposit_df["입금자키"] = deposit_df["입금자(실제)"].astype(str).str.replace(" ", "").str.strip()

        deposit_grouped = deposit_df.groupby("입금자키", as_index=False).agg({
            "입금자(실제)": "first",
            "통장입금": "sum"
        })

        # ✅ 병합 처리
        matched_rows = []
        used_deposit_keys = set()

        for _, order_row in order_grouped.iterrows():
            site_key = order_row["입금자키"]
            matched = False
            for _, deposit_row in deposit_grouped.iterrows():
                deposit_key = deposit_row["입금자키"]
                if (site_key in deposit_key or deposit_key in site_key) and deposit_key not in used_deposit_keys:
                    matched_rows.append({
                        "주문자": order_row["주문자"],
                        "입금자(사이트)": order_row["입금자(사이트)"],
                        "입금자(실제)": deposit_row["입금자(실제)"],
                        "총 구매금액": order_row["총 구매금액"],
                        "통장입금": deposit_row["통장입금"]
                    })
                    used_deposit_keys.add(deposit_key)
                    matched = True
                    break
            if not matched:
                matched_rows.append({
                    "주문자": order_row["주문자"],
                    "입금자(사이트)": order_row["입금자(사이트)"],
                    "입금자(실제)": "",
                    "총 구매금액": order_row["총 구매금액"],
                    "통장입금": 0
                })

        unmatched = deposit_grouped[~deposit_grouped["입금자키"].isin(used_deposit_keys)]
        for _, row in unmatched.iterrows():
            matched_rows.append({
                "주문자": "",
                "입금자(사이트)": "",
                "입금자(실제)": row["입금자(실제)"],
                "총 구매금액": 0,
                "통장입금": row["통장입금"]
            })

        # ✅ 최종 계산
        result_df = pd.DataFrame(matched_rows)
        result_df["차이"] = result_df["통장입금"] - result_df["총 구매금액"]
        result_df = result_df[["주문자", "입금자(사이트)", "입금자(실제)", "총 구매금액", "통장입금", "차이"]].sort_values(by="주문자")

        df_b2b = result_df[~((result_df["주문자"] == "") & (result_df["입금자(사이트)"] == ""))].copy()
        df_non_b2b = result_df[(result_df["주문자"] == "") & (result_df["입금자(사이트)"] == "")].copy()
        df_more_paid = df_b2b[df_b2b["차이"] > 0].copy()
        df_less_paid = df_b2b[df_b2b["차이"] < 0].copy()

        st.success("✅ 정산표가 성공적으로 생성되었습니다!")
        st.dataframe(result_df, use_container_width=True)

        # ✅ 엑셀로 저장 및 강조
        towrite = io.BytesIO()
        with pd.ExcelWriter(towrite, engine="openpyxl") as writer:
            df_b2b.to_excel(writer, index=False, sheet_name="B2B")
            df_non_b2b.to_excel(writer, index=False, sheet_name="B2B 이외")
            df_more_paid.to_excel(writer, index=False, sheet_name="B2B_더 입금된 건들")
            df_less_paid.to_excel(writer, index=False, sheet_name="B2B_덜 입금된 건들")

            workbook = writer.book
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            red_font = Font(color="FF0000", bold=True)
            bold_font = Font(bold=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row):
                    diff = row[5].value
                    if diff is None:
                        continue
                    if sheet_name == "B2B_더 입금된 건들" and diff > 0:
                        row[0].font = bold_font
                    elif sheet_name == "B2B_덜 입금된 건들" and diff < 0:
                        row[0].font = red_font

                    if diff > 0:
                        row[5].fill = yellow_fill
                        row[5].font = bold_font
                    elif diff < 0:
                        row[5].font = red_font

        st.download_button("📥 정산 결과 다운로드", towrite.getvalue(), file_name="정산결과(강조완료).xlsx")

    except Exception as e:
        st.error(f"❌ 오류 발생: {e}")
